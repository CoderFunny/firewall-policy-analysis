# coding=utf-8

import os
import xlrd
import logging
from openpyxl import load_workbook
import re
import IPy
import xlwt
from xlsxwriter.workbook import Workbook
from tqdm import tqdm
import time, datetime

logging.basicConfig(filename='mylog.txt', format="%(asctime)s : %(message)s",
                    level=logging.DEBUG)

protoacolDict = {'TCP': 'ftp,ftp-data,https,smtp,ssh,tcp,qq', 'UDP': 'dns,ntp,snmp,snmptrap,udp,qq'}


# 获取当前路径下所有xls和xlsx文件
def XLSFileList():
    filelist = []
    for root, dirs, files in os.walk(".", topdown=False):
        for name in files:
            str = os.path.join(root, name)
            if '5GC防火墙互通矩阵' in str and str.split('.')[-1] == 'xls' or str.split('.')[-1] == 'xlsx':
                filelist.append(str)
    return filelist


def formatIPData(data):
    dataList = []
    # 临时方案：
    if len(data) == 1 and data[0].count('.') < 3:
        dataList.append('0.0.0.0/0')
    for d in data:
        tmp = re.split('[(（]', d)[0]
        if '.' in tmp and '/' in tmp or ':' in tmp and '/' in tmp:
            dataList.append(tmp)
        if tmp.count('.') == 3 and '/' not in tmp:
            dataList.append(tmp + '/32')

    return dataList


def formatPortData(data):
    dataList = []
    if 'TCP' in data or 'UDP' in data:
        for d in data.replace(" ", "").split('\n'):
            if len(d) and ':' in d or '：' in d:
                portdic = {}
                dk = str(re.split('[:：]', d)[0])
                dv = str(re.split('[:：]', d)[1])
                portdic[dk] = dv
                dataList.append(portdic)
    elif 'TCP' not in data and 'UDP' not in data and '\n' in data:
        dataList.append(data.replace(' ', '').replace('\n', ','))
    else:
        dataList.append(data)
    return dataList


def formatProtocolData(data):
    return re.split('[\n、/]', data)


# IP匹配计算，其中dsip为txt文档中的数据，dtip为excel标准ip
def IPCalculate(dsIP, dtIPList):
    result = False
    for dtIP in dtIPList:
        if '-' not in dtIP:
            if dsIP in IPy.IP(dtIP, make_net=1):
                result = True
                break
        else:
            # 172.20.11.20/30-172.20.11.63/30
            startIPIndex = int(dtIP.split('-')[0].split('/')[0].strip().split('.')[-1])
            endIP = int(dtIP.split('-')[1].split('/')[0].split('.')[-1])
            # print(dtIP)
            for i in range(startIPIndex, endIP + 1):
                IPValue = '.'.join(dtIP.split('-')[0].split('.')[0:3]) + '.' + str(i) + '/' + \
                          dtIP.split('-')[0].split('/')[1]
                if dsIP in IPy.IP(IPValue, make_net=1):
                    result = True
                    break
    return result


def portCalculate(ds, dt, dprotoacol):
    result = False
    for dport in dt:
        if isinstance(dport, str):
            for d1 in re.split('[，,、]', dport):
                if '-' in d1 and '(' not in d1:
                    if int(ds) >= int(d1.split('-')[0]) and int(ds) <= int(d1.split('-')[1].strip(',')):
                        # print(ds, dt)
                        result = True
                        break
                else:
                    if len(d1) and d1.replace('.', '').strip().isdigit():
                        if float(ds) == float(re.split('[（(]', d1)[0]):
                            # print(ds, dt)
                            result = True
                            break
            break
        else:
            for kp in protoacolDict:
                if dprotoacol.upper() in protoacolDict[kp].upper():
                    for key in dport:
                        if key == kp:

                            for d1 in re.split('[，,、]', dport[key]):
                                if '-' in d1 and '(' not in d1:
                                    # print(ds, d1)
                                    if int(ds) >= int(d1.split('-')[0]) and int(ds) <= int(d1.split('-')[1].strip(',')):
                                        # print(ds, dt)
                                        result = True
                                        break
                                else:
                                    if len(d1) and d1.isdigit():
                                        if float(ds) == float(re.split('[（(]', d1)[0]):
                                            # print(ds, dt)
                                            result = True
                                            break
    return result


def VRFCalculate(dsVRF, dtVRF):
    if dsVRF == dtVRF:
        # print(dsVRF, dtVRF)
        return True
    return False


def protoacolCalculate(dsProtoacol, dtProtoacol):
    # for kp in protoacolDict:
    #     if dsProtoacol.upper() in protoacolDict[kp].upper():
    #         protoacoltxt = kp
    #         break
    if dsProtoacol in dtProtoacol:
        return True
    return False


def mappingProtoacol(dsProtoacol):
    protoacol = ''
    for kp in protoacolDict:
        if dsProtoacol.upper() in protoacolDict[kp].upper():
            protoacol = kp
            break
    return protoacol


def FirewallPolicyList(XLSPath):
    # 打开一个workbook
    workbook = xlrd.open_workbook(XLSPath)
    # 抓取所有sheet页的名称
    worksheets = workbook.sheet_names()
    # 定位到目标sheet
    for ws in worksheets:
        if '科学城' in XLSPath and '科学城' in ws:
            worksheet = workbook.sheet_by_name(ws)
            break
        if '白云北' in XLSPath and '白云北' in ws:
            worksheet = workbook.sheet_by_name(ws)
            break
    # 获取该sheet中的有效行数
    num_rows = worksheet.nrows
    # 获取列表的有效列数
    num_cols = worksheet.ncols

    # 打开excel
    wb = load_workbook(XLSPath)
    # 获取工作表
    # for name in wb.sheetnames:
    #     if '广东科学城DC1' in name:
    #         sheet = wb[name]

    logging.info('read %s start.', XLSPath)
    # 解析xls
    destColnDict = {'源VRF列': 0, '源网段列': 0, '源端口列': 0, '目的VRF列': 0, '目的网段列': 0, '目的端口列': 0, '协议列': 0}
    StandardDataList = []

    # 记录各字段所在列
    for coln in range(num_cols):
        if '源VRF' in str(worksheet.cell_value(0, coln)):
            destColnDict['源VRF列'] = coln
        if '源网段' in str(worksheet.cell_value(0, coln)):
            destColnDict['源网段列'] = coln
        if '源端口' in str(worksheet.cell_value(0, coln)):
            destColnDict['源端口列'] = coln
        if '目的VRF' in str(worksheet.cell_value(0, coln)):
            destColnDict['目的VRF列'] = coln
        if '目的网段' in str(worksheet.cell_value(0, coln)):
            destColnDict['目的网段列'] = coln
        if '目的端口' in str(worksheet.cell_value(0, coln)):
            destColnDict['目的端口列'] = coln
        if '协议' in str(worksheet.cell_value(0, coln)):
            destColnDict['协议列'] = coln
    for rown in range(1, num_rows):
        DataDict = {}
        DataDict['源VRF(标准)'] = str(worksheet.cell_value(rown, destColnDict['源VRF列'])).strip()
        DataDict['目的VRF(标准)'] = str(worksheet.cell_value(rown, destColnDict['目的VRF列'])).strip()
        DataDict['源网段(标准)'] = formatIPData(str(worksheet.cell_value(rown, destColnDict['源网段列'])).split('\n'))
        DataDict['目的网段(标准)'] = formatIPData(str(worksheet.cell_value(rown, destColnDict['目的网段列'])).split('\n'))
        DataDict['源端口(标准)'] = formatPortData(str(worksheet.cell_value(rown, destColnDict['源端口列'])))
        DataDict['目的端口(标准)'] = formatPortData(str(worksheet.cell_value(rown, destColnDict['目的端口列'])))
        DataDict['协议(标准)'] = formatProtocolData(str(worksheet.cell_value(rown, destColnDict['协议列'])))
        StandardDataList.append(DataDict)
    # for i in StandardDataList:
    #     print(i)
    logging.info('write %s end.', XLSPath)
    # testAccuracy(StandardDataList)
    return StandardDataList


def string_duplicate_1(s):
    new_s = list(set(s))  # set无序
    new_s.sort(key=s.index)
    return new_s


def testAccuracy(list):
    # 写数据时，行计数器
    logging.info('xls write begin')
    # 实例化一个execl对象xls=工作薄
    xls = xlwt.Workbook()
    # 实例化一个工作表，名叫Sheet1
    sht1 = xls.add_sheet(u'test')
    # 第一个参数是行，第二个参数是列，第三个参数是值,第四个参数是格式
    headFont = SetFont(1)
    bodyFont1 = SetFont(4)  # 水平垂直居中
    bodyFont2 = SetFont(5)
    bodyFont3 = SetFont(6)
    bodyFont4 = SetFont(2)
    bodyFont5 = SetFont(3)
    bodyFont6 = SetFont(7)
    bodyFont7 = SetFont(8)
    sht1.write(0, 0, '源VRF(实际)', headFont)
    sht1.write(0, 1, '源网段(实际)', headFont)
    sht1.write(0, 2, '源端口(实际)', headFont)
    sht1.write(0, 3, '目的VRF(实际)', headFont)
    sht1.write(0, 4, '目的网段(实际)', headFont)
    sht1.write(0, 5, '目的端口(实际)', headFont)
    sht1.write(0, 6, '协议(实际)', headFont)

    # 数据写入
    # sheet1
    shtNum1 = 1

    matchXlsList = list
    sht1.col(1).width = 256 * 20
    sht1.col(2).width = 256 * 20
    sht1.col(4).width = 256 * 20
    for i in range(len(matchXlsList)):
        sht1.write(shtNum1, 0, matchXlsList[i]['源VRF'], bodyFont2)
        sht1.write(shtNum1, 1, '\n'.join(matchXlsList[i]['源网段']), bodyFont2)
        sht1.write(shtNum1, 2, matchXlsList[i]['源端口'], bodyFont7)
        sht1.write(shtNum1, 3, matchXlsList[i]['目的VRF'], bodyFont2)
        sht1.write(shtNum1, 4, '\n'.join(matchXlsList[i]['目的网段']), bodyFont2)
        if isinstance(matchXlsList[i]['目的端口'][0], str):
            sht1.write(shtNum1, 5, matchXlsList[i]['目的端口'], bodyFont7)
        else:
            strtmp = ''
            for dic in matchXlsList[i]['目的端口']:
                for key in dic:
                    strtmp = strtmp + key + ':' + dic[key] + '\n'
            sht1.write(shtNum1, 5, strtmp, bodyFont2)
        sht1.write(shtNum1, 6, '\n'.join(matchXlsList[i]['协议']), bodyFont7)
        shtNum1 = shtNum1 + 1

    xls.save(os.getcwd() + '\\DC1科学城' + '\\FirewallPolicyResult-kxc' + '.xls')
    logging.info('xls write end')


def mergeCondition(fl, slist):
    result = False
    # print(fl['源VRF'] == sl['源VRF'])
    # print(fl['源网段'] == sl['源网段'])
    # print(fl['目的VRF'] == sl['目的VRF'])
    # print(fl['目的网段'] == sl['目的网段'])
    # print(fl['目的端口'] == sl['目的网段'])
    # print(fl['映射协议'] == sl['映射协议'])
    # print(int(fl['源端口']) >= 1000)
    # print(int(sl['源端口']) >= 1000)
    # print('\n')
    if isinstance(slist, dict):
        sl = slist
        if fl['源VRF'] + fl['源网段'] + fl['目的VRF'] + fl['目的网段'] + fl['目的端口'] + fl['映射协议'] == \
                sl['源VRF'] + sl['源网段'] + sl['目的VRF'] + sl['目的网段'] + sl['目的端口'] + sl['映射协议'] and \
                int(fl['源端口']) >= 1000 and int(sl['源端口']) >= 1000:
            result = True
    if isinstance(slist, list) and len(slist):
        for sl in slist:
            if fl['源VRF'] + fl['源网段'] + fl['目的VRF'] + fl['目的网段'] + fl['目的端口'] + fl['映射协议'] == \
                    sl['源VRF'] + sl['源网段'] + sl['目的VRF'] + sl['目的网段'] + sl['目的端口'] + sl['映射协议'] and \
                    int(fl['源端口']) >= 1000 and int(sl['源端口']) >= 1000:
                result = True
    return result


def mergeDictData(matchList, mismatchList):
    totalList = matchList + mismatchList
    resultList = []
    for i in range(len(totalList)):
        if not mergeCondition(totalList[i], resultList):
            resultList.append(totalList[i])
    # for i in resultList:
    #     print(i)
    return resultList


def FirewallPolicyAnalysis(LogPath, standardList, pbar):
    logging.info('begin to analysis txt file,%s', LogPath)
    # try:
    dataLists = []
    dataList = []
    sourceDataList = []
    file = open(LogPath, "r", encoding='utf-16-le')
    for line in file.readlines():
        dataLists.append(line.strip('\n\t\r').replace('"', ''))

    for dts in dataLists:
        if '-->' in dts and dts.count('-->') == 2:
            dataList.append(dts)
        if '-->' in dts and '+->' in dts:
            dataList.append(dts.replace('+->', '-->'))
    dataList = string_duplicate_1(dataList)
    for dt in dataList:
        dict = {'源VRF': '', '目的VRF': '', '源网段': '', '源端口': '', '目的网段': '', '目的端口': '', '协议': '', '映射协议': ''}
        # d = dt.replace('Remote', '').split('-->')
        d = ' '.join(re.sub(' +', ' ', dt.replace('Remote', '')).split(' ')).split('-->')
        dict['协议'] = d[0].split(':')[0].strip(' ').split(' ')[0].lstrip('\ufeff')
        dict['映射协议'] = mappingProtoacol(dict['协议'])
        dict['源VRF'] = d[0].split(':')[1].strip(' ')
        dict['目的VRF'] = d[1].strip(' ').split(' ')[0]
        if d[1].strip(' ').split(' ')[1].count('.') == 3:
            dict['源网段'] = d[1].strip(' ').split(' ')[1].split(':')[0]
            dict['源端口'] = d[1].strip(' ').split(' ')[1].split(':')[1]
        if d[1].strip(' ').split(' ')[1].count(':') > 1:
            dict['源网段'] = d[1].strip(' ').split(' ')[1].split('.')[0]
            dict['源端口'] = d[1].strip(' ').split(' ')[1].split('.')[1]
        if d[2].strip(' ').count('.') == 3:
            dict['目的网段'] = d[2].strip(' ').split(':')[0]
            dict['目的端口'] = d[2].strip(' ').split(':')[1].split('(')[0]
        if d[2].strip(' ').count(':') > 1:
            dict['目的网段'] = d[2].strip(' ').split('.')[0]
            dict['目的端口'] = d[2].strip(' ').split('.')[1].split('(')[0]
        sourceDataList.append(dict)
    matchTxtList = []
    mismatchTxtList = []
    resultDict = {}
    count = 0
    resultList = []
    for sl in sourceDataList:
        count += 1
        matchFlag = 0
        if count % 100 == 0:
            pbar.update(50 * 100 / len(sourceDataList))
        for tl in standardList:
            # print(portCalculate(sl['源端口'], tl['源端口(标准)'], sl['协议']))
            # print(portCalculate(sl['目的端口'], tl['目的端口(标准)'], sl['协议']))
            # print(IPCalculate(sl['源网段'], tl['源网段(标准)']))
            # print(IPCalculate(sl['目的网段'], tl['目的网段(标准)']))
            # print(VRFCalculate(sl['源VRF'], tl['源VRF(标准)']))
            # print(VRFCalculate(sl['目的VRF'], tl['目的VRF(标准)']))
            # print(protoacolCalculate(sl['映射协议'], tl['协议(标准)']))
            if portCalculate(sl['源端口'], tl['源端口(标准)'], sl['协议']) and portCalculate(sl['目的端口'], tl['目的端口(标准)'], sl['协议']) \
                    and IPCalculate(sl['源网段'], tl['源网段(标准)']) and IPCalculate(sl['目的网段'], tl['目的网段(标准)']) \
                    and VRFCalculate(sl['源VRF'], tl['源VRF(标准)']) and VRFCalculate(sl['目的VRF'], tl['目的VRF(标准)']) \
                    and protoacolCalculate(sl['映射协议'], tl['协议(标准)']):
                matchTxtList.append({**sl, **tl})
                matchFlag = 1
                # if not mergeCondition(sl, resultList):
                #     resultList.append({**sl, **tl})
                break
        if matchFlag == 0:
            mismatchTxtList.append(sl)
            # if not mergeCondition(sl, resultList):
            #     resultList.append(sl)
    # for i in resultList:
    #     print(i)
    # print(len(resultList))
    resultDict['initialResult'] = matchTxtList + mismatchTxtList
    # 实际值中如果其余六项相同，源端口号大于1000的，数据合并
    resultDict['mergeResult'] = mergeDictData(matchTxtList, mismatchTxtList)
    # resultDict['tmp'] = resultList

    file.close()
    # except Exception as err:
    #     logging.error('txtAnalysis function error : %s', err)
    # finally:
    #     if file:
    #         file.close()
    logging.info('end analysis mml file,%s', LogPath)
    return resultDict


# 设置单元格式 入参type (1:表头第一列样式  2:某一单元格样式)
def SetFont(type):
    style = xlwt.XFStyle()
    pattern = xlwt.Pattern()
    borders = xlwt.Borders()
    al = xlwt.Alignment()
    # 设置边框
    borders.left = 1
    borders.right = 1
    borders.top = 1
    borders.bottom = 1
    borders.bottom_colour = 0x3A
    style.borders = borders

    if type == 1:
        # 设置字体格式
        Font = xlwt.Font()
        Font.name = "Times New Roman"
        Font.bold = True  # 加粗
        style.font = Font

        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al
        style.alignment.wrap = 1

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['yellow']
        style.pattern = pattern
    elif type == 2:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x00  # 设置垂直居中
        style.alignment = al
        style.alignment.wrap = 1

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['light_green']
        style.pattern = pattern
    elif type == 3:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x00  # 设置垂直居中
        style.alignment = al
        style.alignment.wrap = 1

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['sky_blue']
        style.pattern = pattern
    elif type == 4:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al
        style.alignment.wrap = 1

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['light_green']
        style.pattern = pattern
    elif type == 5:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al
        style.alignment.wrap = 1

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['sky_blue']
        style.pattern = pattern
    elif type == 6:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al
        style.alignment.wrap = 1

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['light_yellow']
        style.pattern = pattern
    elif type == 7:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['light_green']
        style.pattern = pattern
    elif type == 8:
        # 水平垂直居中
        al.horz = 0x02  # 设置水平居中
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al

        # 设置单元格背景色
        pattern.pattern = xlwt.Pattern.SOLID_PATTERN
        pattern.pattern_fore_colour = xlwt.Style.colour_map['sky_blue']
        style.pattern = pattern
    return style


def writeSheet(sheet, resultList):
    headFont = SetFont(1)
    bodyFont1 = SetFont(4)  # 水平垂直居中light_green
    bodyFont2 = SetFont(5)  # 水平垂直居中sky_blue
    bodyFont3 = SetFont(6)  # 水平垂直居中light_yellow
    bodyFont6 = SetFont(7)  # 水平垂直居中不换行light_green
    bodyFont7 = SetFont(8)  # 水平垂直居中不换行sky_blue
    sheet.write(0, 0, '源VRF(实际)', headFont)
    sheet.write(0, 1, '源网段(实际)', headFont)
    sheet.write(0, 2, '源端口(实际)', headFont)
    sheet.write(0, 3, '目的VRF(实际)', headFont)
    sheet.write(0, 4, '目的网段(实际)', headFont)
    sheet.write(0, 5, '目的端口(实际)', headFont)
    sheet.write(0, 6, '协议(实际)', headFont)
    sheet.write(0, 7, '映射协议', headFont)
    sheet.write(0, 8, '源VRF(标准)', headFont)
    sheet.write(0, 9, '源网段(标准)', headFont)
    sheet.write(0, 10, '源端口(标准)', headFont)
    sheet.write(0, 11, '目的VRF(标准)', headFont)
    sheet.write(0, 12, '目的网段(标准)', headFont)
    sheet.write(0, 13, '目的端口(标准)', headFont)
    sheet.write(0, 14, '协议(标准)', headFont)
    sheet.write(0, 15, '分析结果', headFont)

    sheet.col(1).width = 256 * 13
    sheet.col(4).width = 256 * 13
    sheet.col(9).width = 256 * 15
    sheet.col(12).width = 256 * 15
    sheet.col(13).width = 256 * 26
    # 数据写入
    shtNum1 = 1
    if len(resultList):
        for matchData in resultList:
            if len(matchData) == 15:
                sheet.write(shtNum1, 0, matchData['源VRF'], bodyFont1)
                sheet.write(shtNum1, 1, matchData['源网段'], bodyFont1)
                sheet.write(shtNum1, 2, matchData['源端口'], bodyFont6)
                sheet.write(shtNum1, 3, matchData['目的VRF'], bodyFont1)
                sheet.write(shtNum1, 4, matchData['目的网段'], bodyFont1)
                sheet.write(shtNum1, 5, matchData['目的端口'], bodyFont6)
                sheet.write(shtNum1, 6, matchData['协议'], bodyFont1)
                sheet.write(shtNum1, 7, matchData['映射协议'], bodyFont1)

                sheet.write(shtNum1, 8, matchData['源VRF(标准)'], bodyFont2)
                sheet.write(shtNum1, 9, '\n'.join(matchData['源网段(标准)']), bodyFont2)
                sheet.write(shtNum1, 10, matchData['源端口(标准)'], bodyFont7)
                sheet.write(shtNum1, 11, matchData['目的VRF(标准)'], bodyFont2)
                sheet.write(shtNum1, 12, '\n'.join(matchData['目的网段(标准)']), bodyFont2)
                if isinstance(matchData['目的端口(标准)'][0], str):
                    sheet.write(shtNum1, 13, matchData['目的端口(标准)'], bodyFont7)
                else:
                    strtmp = ''
                    for dic in matchData['目的端口(标准)']:
                        for key in dic:
                            strtmp = strtmp + key + ':' + dic[key] + '\n'
                    sheet.write(shtNum1, 13, strtmp, bodyFont2)
                sheet.write(shtNum1, 14, '\n'.join(matchData['协议(标准)']), bodyFont2)
                sheet.write(shtNum1, 15, 'PASS', bodyFont3)
                shtNum1 = shtNum1 + 1

            if len(matchData) == 8:
                sheet.write(shtNum1, 0, matchData['源VRF'], bodyFont1)
                sheet.write(shtNum1, 1, matchData['源网段'], bodyFont1)
                sheet.write(shtNum1, 2, matchData['源端口'], bodyFont6)
                sheet.write(shtNum1, 3, matchData['目的VRF'], bodyFont1)
                sheet.write(shtNum1, 4, matchData['目的网段'], bodyFont1)
                sheet.write(shtNum1, 5, matchData['目的端口'], bodyFont6)
                sheet.write(shtNum1, 6, matchData['协议'], bodyFont1)
                sheet.write(shtNum1, 7, matchData['映射协议'], bodyFont1)
                shtNum1 = shtNum1 + 1


def XLSWrite(XLSPath, FPDict):
    # 写数据时，行计数器
    logging.info('xls write begin')
    # 实例化一个execl对象xls=工作薄
    xls = xlwt.Workbook()
    # 实例化一个工作表，名叫Sheet1
    sht1 = xls.add_sheet(u'防火墙配置策略分析结果')
    sht2 = xls.add_sheet(u'防火墙配置策略分析结果2')
    writeSheet(sht1, FPDict['initialResult'])
    writeSheet(sht2, FPDict['mergeResult'])
    # writeSheet(sht3, FPDict['tmp'])
    # # 第一个参数是行，第二个参数是列，第三个参数是值,第四个参数是格式
    # headFont = SetFont(1)
    # bodyFont1 = SetFont(4)  # 水平垂直居中light_green
    # bodyFont2 = SetFont(5)  # 水平垂直居中sky_blue
    # bodyFont3 = SetFont(6)  # 水平垂直居中light_yellow
    # bodyFont4 = SetFont(2)
    # bodyFont5 = SetFont(3)
    # bodyFont6 = SetFont(7)  # 水平垂直居中不换行light_green
    # bodyFont7 = SetFont(8)  # 水平垂直居中不换行sky_blue
    # sht1.write(0, 0, '源VRF(实际)', headFont)
    # sht1.write(0, 1, '源网段(实际)', headFont)
    # sht1.write(0, 2, '源端口(实际)', headFont)
    # sht1.write(0, 3, '目的VRF(实际)', headFont)
    # sht1.write(0, 4, '目的网段(实际)', headFont)
    # sht1.write(0, 5, '目的端口(实际)', headFont)
    # sht1.write(0, 6, '协议(实际)', headFont)
    # sht1.write(0, 7, '映射协议', headFont)
    # sht1.write(0, 8, '源VRF(标准)', headFont)
    # sht1.write(0, 9, '源网段(标准)', headFont)
    # sht1.write(0, 10, '源端口(标准)', headFont)
    # sht1.write(0, 11, '目的VRF(标准)', headFont)
    # sht1.write(0, 12, '目的网段(标准)', headFont)
    # sht1.write(0, 13, '目的端口(标准)', headFont)
    # sht1.write(0, 14, '协议(标准)', headFont)
    # sht1.write(0, 15, '分析结果', headFont)
    #
    # # 数据写入
    # # sheet1
    # shtNum1 = 1
    # if len(FPDict['matchTxt']):
    #     matchTxtList = FPDict['matchTxt']
    #
    #     sht1.col(1).width = 256 * 13
    #     sht1.col(4).width = 256 * 13
    #     sht1.col(9).width = 256 * 15
    #     sht1.col(12).width = 256 * 15
    #     sht1.col(13).width = 256 * 26
    #     for matchData in matchTxtList:
    #         sht1.write(shtNum1, 0, matchData['源VRF'], bodyFont1)
    #         sht1.write(shtNum1, 1, matchData['源网段'], bodyFont1)
    #         sht1.write(shtNum1, 2, matchData['源端口'], bodyFont6)
    #         sht1.write(shtNum1, 3, matchData['目的VRF'], bodyFont1)
    #         sht1.write(shtNum1, 4, matchData['目的网段'], bodyFont1)
    #         sht1.write(shtNum1, 5, matchData['目的端口'], bodyFont6)
    #         sht1.write(shtNum1, 6, matchData['协议'], bodyFont1)
    #         sht1.write(shtNum1, 7, matchData['映射协议'], bodyFont1)
    #
    #         sht1.write(shtNum1, 8, matchData['源VRF(标准)'], bodyFont2)
    #         sht1.write(shtNum1, 9, '\n'.join(matchData['源网段(标准)']), bodyFont2)
    #         sht1.write(shtNum1, 10, matchData['源端口(标准)'], bodyFont7)
    #         sht1.write(shtNum1, 11, matchData['目的VRF(标准)'], bodyFont2)
    #         sht1.write(shtNum1, 12, '\n'.join(matchData['目的网段(标准)']), bodyFont2)
    #         if isinstance(matchData['目的端口(标准)'][0], str):
    #             sht1.write(shtNum1, 13, matchData['目的端口(标准)'], bodyFont7)
    #         else:
    #             strtmp = ''
    #             for dic in matchData['目的端口(标准)']:
    #                 for key in dic:
    #                     strtmp = strtmp + key + ':' + dic[key] + '\n'
    #             sht1.write(shtNum1, 13, strtmp, bodyFont2)
    #         sht1.write(shtNum1, 14, '\n'.join(matchData['协议(标准)']), bodyFont2)
    #         sht1.write(shtNum1, 15, 'PASS', bodyFont3)
    #         shtNum1 = shtNum1 + 1
    #
    # if len(FPDict['mismatch']):
    #     mismatchList = FPDict['mismatch']
    #     for mismatchData in mismatchList:
    #         sht1.write(shtNum1, 0, mismatchData['源VRF'], bodyFont1)
    #         sht1.write(shtNum1, 1, mismatchData['源网段'], bodyFont1)
    #         sht1.write(shtNum1, 2, mismatchData['源端口'], bodyFont6)
    #         sht1.write(shtNum1, 3, mismatchData['目的VRF'], bodyFont1)
    #         sht1.write(shtNum1, 4, mismatchData['目的网段'], bodyFont1)
    #         sht1.write(shtNum1, 5, mismatchData['目的端口'], bodyFont6)
    #         sht1.write(shtNum1, 6, mismatchData['协议'], bodyFont1)
    #         sht1.write(shtNum1, 7, mismatchData['映射协议'], bodyFont1)
    #         shtNum1 = shtNum1 + 1
    xls.save(XLSPath)
    logging.info('xls write end')


def xlsxWriter():
    # 创建Excel对象
    workbook = Workbook('a.xlsx')
    worksheet = workbook.add_worksheet()
    color = workbook.add_format({'color': 'red', 'bold': True})

    # 日期高亮
    rc = re.compile('([0-9年月日]{2,})')
    sentence = '小洪和小黄2020年1月12日母校\n初见。1月26日长烟落日孤城闭，2月9日神仙眷侣云比心'
    format_ls = rc.split(sentence)
    for i in range(len(format_ls) - 1, -1, -1):
        if rc.fullmatch(format_ls[i]):
            format_ls.insert(i, color)  # Prefix the word with the format
    print(format_ls)

    # 写入单元格
    row, col = 2, 1
    worksheet.write(row, col, '\n'.join(format_ls))
    # worksheet.write_rich_string(row, col, *format_ls)
    workbook.close()


def fileCount(kxcpath, bybpath):
    kxcFileCount = 0
    bybFileCount = 0
    for f in os.listdir(kxcpath):
        if '防火墙' in f and f.endswith('xlsx'):
            kxcFileCount += 1
    for f in os.listdir(kxcpath):
        if f.endswith('.log'):
            kxcFileCount += 1

    for f in os.listdir(bybpath):
        if '防火墙' in f and f.endswith('xlsx'):
            bybFileCount += 1
    for f in os.listdir(bybpath):
        if f.endswith('.log'):
            bybFileCount += 1
    return kxcFileCount, bybFileCount


def main():
    # 解析xls文件到list，用于后续数据处理数据源
    logging.info('welcome to firewall polocy analysis world.')
    kxcpath = os.getcwd() + '\\DC1科学城'
    bybpath = os.getcwd() + '\\DC2白云北'

    # try:
    FPkxcList = []
    FPkxcDict = {}

    for i in range(0, 20):
        kongxin = chr(9711)
        shixin = chr(9679) + kongxin
        print(shixin, end="")
    print('\n')
    print('【************  防火墙策略数据分析中，请勿关闭窗口  ************】\n')
    kxcFileCount, bybFileCount = fileCount(kxcpath, bybpath)
    if kxcFileCount >= 2 and bybFileCount >= 2:
        pbar = tqdm(total=100, ncols=50, bar_format='当前进度:{percentage:3.0f}%|{bar}|数据处理中...')
    if kxcFileCount >= 2 and bybFileCount < 2 or kxcFileCount < 2 and bybFileCount >= 2:
        pbar = tqdm(total=50, ncols=50, bar_format='当前进度:{percentage:3.0f}%|{bar}|{n}/{total}')
    if kxcFileCount + bybFileCount == 0:
        pbar = tqdm(total=0, ncols=50, bar_format='当前进度:{percentage:3.0f}%|{bar}|{n}/{total}')
        print('文件不存在，请检查')

    for f in os.listdir(kxcpath):
        if '防火墙' in f and f.endswith('xlsx'):
            FPkxcList = FirewallPolicyList(kxcpath + '\\' + f)
    for f in os.listdir(kxcpath):
        if f.endswith('.log') and len(FPkxcList):
            FPkxcDict = FirewallPolicyAnalysis(kxcpath + '\\' + f, FPkxcList, pbar)
    if len(FPkxcDict):
        date = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        XLSWrite(kxcpath + '\\FirewallPolicyResult-kxc-' + str(date) + '.xls', FPkxcDict)
    # 白云北
    FPbybList = []
    FPbybDict = {}
    for f in os.listdir(bybpath):
        if '防火墙' in f and f.endswith('xlsx'):
            FPbybList = FirewallPolicyList(bybpath + '\\' + f)
    for f in os.listdir(bybpath):
        if f.endswith('.log') and len(FPbybList):
            FPbybDict = FirewallPolicyAnalysis(bybpath + '\\' + f, FPbybList, pbar)
    if len(FPbybDict):
        date = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        XLSWrite(bybpath + '\\FirewallPolicyResult-byb-' + str(date) + '.xls', FPbybDict)
    # except Exception as err:
    #     logging.error(err)
    pbar.close()
    print('\n[************  已完成分析  ************]\n')
    for i in range(0, 20):
        kongxin = chr(9711)
        shixin = chr(9679) + kongxin
        print(shixin, end="")
    logging.info("end firewall polocy analysis world")


if __name__ == '__main__':
    # print(time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time())))
    # print(datetime.date.today())
    main()
